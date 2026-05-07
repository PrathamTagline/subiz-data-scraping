import json
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


def normalize(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def sheet_to_records(sheet):
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return []

    headers = [
        (str(h).strip() if h is not None else f"column_{i + 1}")
        for i, h in enumerate(headers)
    ]

    records = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        records.append({h: normalize(c) for h, c in zip(headers, row)})
    return records


def workbook_to_dict(path: Path) -> dict:
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    return {sheet.title: sheet_to_records(sheet) for sheet in wb.worksheets}


def convert_file(src: Path, dst_dir: Path) -> Path:
    data = workbook_to_dict(src)
    dst_dir.mkdir(parents=True, exist_ok=True)
    out = dst_dir / (src.stem + ".json")
    with out.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out


def convert_directory(src_dir: Path, dst_dir: Path):
    results = []
    for src in sorted(src_dir.glob("*.xlsx")):
        out = convert_file(src, dst_dir)
        results.append((src, out))
    return results
